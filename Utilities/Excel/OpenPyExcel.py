# -*- coding: utf-8 -*-
from PyQt5.QtCore import center
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class PersonalExcel():
    def __init__(self, filename = "", sheetname = ""):
        super(PersonalExcel, self).__init__()
        # 添加属性
        self.filename  = filename
        self.sheetname = sheetname
        self.workbook  = None
        self.worksheet = None

    def initWorkBook(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname
        self.workbook = Workbook(filename) # 创建工作簿
        self.worksheet = self.workbook.create_sheet(self.sheetname, 0)
        self.worksheet.title = self.sheetname
        self.worksheet.sheet_properties.tabColor = "1072BA"
        self.workbook.save(self.filename)

    def closeFile(self):
        self.workbook.close()

    def writeData(self, filename, row, col, val):
        self.workbook = load_workbook(filename)
        self.worksheet = self.workbook.active # 默认工作表
        self.worksheet.cell(row, col, val)
        self.workbook.save(filename)
        self.workbook.close()

    def wrtieRow(self, filename, rowList): # 会覆盖之前的内容
        self.workbook = load_workbook(filename)
        self.worksheet = self.workbook.active # 默认工作表
        self.worksheet.append(rowList)
        self.workbook.save(filename)
        self.workbook.close()

    def saveFile(self):
        self.workbook.save(self.filename)