# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Color, colors, NamedStyle
from openpyxl.styles.fills import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = 'Test'
tfont = Font(name='Times New Roman', size=24, bold=False, color=colors.BLUE)
talignment = Alignment(horizontal='center', vertical='center', textRotation=0, wrapText=False, shrinkToFit=True)
ws1['A1'] = 'Hello'
ws1['A1'].font = tfont
ws1['A1'].alignment = talignment

tfont = Font(name='Times New Roman', size=12, bold=True, color=colors.BLACK)
ws1['A2'] = 'World'
ws1['A2'].font = tfont

ws2 = wb.create_sheet('Formula')
ws2['B1'] = 200
ws2['B2'] = 500
ws2['B3'] = '=SUM(B1:B2)'

ws3 = wb.create_sheet('Dimension')
ws3['A1'] = 'Row Tall'
ws3.row_dimensions[1].height = 700
ws3['B2'] = 'Column Wide'
ws3.column_dimensions['B'].width = 200

ws4 = wb.create_sheet('Merged')
ws4.merge_cells('A1:D3')
ws4['A1'] = 'Merge 12 Cells'
ws4.merge_cells('C5:D5')
ws4['C5'] = 'Merge 12 Cells'

ws5 = wb.copy_worksheet(ws4)
ws5.title = 'Unmerged'
ws5.unmerge_cells('A1:D3')
ws5.unmerge_cells('C5:D5')

ws6 = wb.create_sheet('Named Style')
# 检测结果通过
resultPassStyle = NamedStyle('resultPassStyle')
resultPassStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=0))
resultPassStyle.alignment = Alignment(horizontal='center', vertical='center')
resultPassStyle.fill = PatternFill('solid', '0000CCFF') # 蓝色填充背景
# 检测结果失败
resultFailStyle = NamedStyle('resultFailStyle')
resultFailStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=0))
resultFailStyle.alignment = Alignment(horizontal='center', vertical='center')
resultFailStyle.fill = PatternFill('solid', '00FF0000') # 红色填充背景
# 正常数值或结果
normalStyle = NamedStyle('normalStyle')
normalStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=4))# 蓝色字体
normalStyle.alignment = Alignment(horizontal='center', vertical='center')
# 异常数值或结果
unnormalStyle = NamedStyle('unnormalStyle')
unnormalStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=2)) # 红色字体
unnormalStyle.alignment = Alignment(horizontal='center', vertical='center')

wb.add_named_style(resultPassStyle)
wb.add_named_style(resultFailStyle)
wb.add_named_style(normalStyle)
wb.add_named_style(unnormalStyle)

ws6['A1'] = '通过'
ws6['A1'].style = 'resultPassStyle'
ws6['B1'] = '失败'
ws6['B1'].style = 'resultFailStyle'
ws6['A2'] = '成功'
ws6['A2'].style = 'normalStyle'
ws6['B2'] = '120.3'
ws6['B2'].style = 'unnormalStyle'

wb.save('styles.xlsx')