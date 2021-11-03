# -*- coding: utf-8 -*-
# 全局变量getset
import GetSetObj
# openpyxl相关模块
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, NamedStyle, Side
from openpyxl.styles.fills import PatternFill  # 填充


class PrivateOpenPyxl():
    #*------------------------------------------------命名样式------------------------------------------------*#
    #*---------------------------------------------Named Styles----------------------------------------------*#
    # 检测结果通过
    resultPassStyle = NamedStyle('resultPassStyle')
    resultPassStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=0))# 黑色字体
    resultPassStyle.alignment = Alignment(horizontal='center', vertical='center')
    resultPassStyle.fill = PatternFill('solid', '0000CCFF') # 蓝色填充背景
    resultPassStyle.border = Border(left   = Side('thin', color=Color(indexed=0)),
                                    right  = Side('thin', color=Color(indexed=0)),
                                    top    = Side('thin', color=Color(indexed=0)),
                                    bottom = Side('thin', color=Color(indexed=0)))
    # 检测结果失败
    resultFailStyle = NamedStyle('resultFailStyle')
    resultFailStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=0))# 黑色字体
    resultFailStyle.alignment = Alignment(horizontal='center', vertical='center')
    resultFailStyle.fill = PatternFill('solid', '00FF0000') # 红色填充背景
    resultFailStyle.border = Border(left   = Side('thin', color=Color(indexed=0)),
                                    right  = Side('thin', color=Color(indexed=0)),
                                    top    = Side('thin', color=Color(indexed=0)),
                                    bottom = Side('thin', color=Color(indexed=0)))
    # 默认字体
    defaultContentStyle = NamedStyle('defaultContentStyle')
    defaultContentStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=0))# 黑色字体
    defaultContentStyle.alignment = Alignment(horizontal='center', vertical='center')
    # 正常数值或结果
    normalContentStyle = NamedStyle('normalContentStyle')
    normalContentStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=4))# 蓝色字体
    normalContentStyle.alignment = Alignment(horizontal='center', vertical='center')
    # 异常数值或结果
    abnormalContentStyle = NamedStyle('abnormalContentStyle')
    abnormalContentStyle.font = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=2)) # 红色字体
    abnormalContentStyle.alignment = Alignment(horizontal='center', vertical='center')

    def __init__(self, wbname = "", wsname = ""):
        super(PrivateOpenPyxl, self).__init__()
        # 添加属性
        self.wbname  = wbname # 工作簿名
        self.wsname = wsname  # 工作表名
        self.wb = None        # 工作簿对象
        self.ws = None        # 工作表对象
        #*------------------------------------------------一般样式------------------------------------------------*#
        #*----------------------------------------------Cell Styles----------------------------------------------*#
        # 表头单元格样式
        self.hfont = Font(name='Times New Roman', size=16, bold=False, color=Color(indexed=17))
        self.halignment = Alignment(horizontal='center', vertical='center', textRotation=0, wrapText=False)
        self.hfill = PatternFill('solid', '00FFFF00') # 黄色填充背景
        self.hborder = Border(left   = Side('thin', color=Color(indexed=0)),
                              right  = Side('thin', color=Color(indexed=0)),
                              top    = Side('thin', color=Color(indexed=0)),
                              bottom = Side('thin', color=Color(indexed=0)))
        # 默认列宽 对应字体大小如上
        self.columnWidth = [12, 30, 14.38, 17.25, 9, 11.88, 11.88, 12.13, 10, 11.88, 11.88, 12.13, 10, 11.88, 6.38]
        # 26列列索引
        self.colindex = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        
    def createWorkBook(self, wbname, wsname):
        # 创建工作簿工作表
        self.wbname = wbname
        self.wsname = wsname
        self.wb = Workbook(wbname) # 创建工作簿
        self.ws = self.wb.create_sheet(self.wsname, 0)
        self.ws.title = self.wsname
        self.ws.sheet_properties.tabColor = "1072BA"
        self.wb.add_named_style(self.resultPassStyle)
        self.wb.add_named_style(self.resultFailStyle)
        self.wb.add_named_style(self.defaultContentStyle)
        self.wb.add_named_style(self.normalContentStyle)
        self.wb.add_named_style(self.abnormalContentStyle)
        self.wb.save(self.wbname)

    def addNamedStyle(self):
        # 添加命名样式 样式添加后一定要保存一次!!!
        self.wb.add_named_style(self.resultPassStyle)
        self.wb.add_named_style(self.resultFailStyle)
        self.wb.add_named_style(self.defaultContentStyle)
        self.wb.add_named_style(self.normalContentStyle)
        self.wb.add_named_style(self.abnormalContentStyle)

    def loadSheet(self, wbname):
        self.wbname = wbname
        self.wb = load_workbook(wbname)
        self.ws = self.wb.active # 默认工作表    

    def saveSheet(self):
        self.wb.save(self.wbname)
    
    def closeSheet(self):
        self.wb.close()

    def writeData(self, wbname, row, col, val):
        self.wb = load_workbook(wbname)
        self.ws = self.wb.active # 默认工作表
        self.ws.cell(row, col, val)
        self.wb.save(wbname)

    def appendRow(self, rowList):
        self.ws.append(rowList)

    def deleteRow(self, rowIndex):
        self.ws.delete_rows(rowIndex)

    def insertRow(self, rowIndex):
        self.ws.insert_rows(rowIndex)

    def deleteCloumn(self, colIndex):
        self.ws.delete_cols(colIndex)

    def getRowIndexByID(self, id):
        maxrows = self.ws.max_row # 每次执行行数据填入都获取最新的最大行（包括表头所在行）
        maxcols = self.ws.max_column # 每次执行行数据填入都获取最新的最大列
        gen = self.ws.iter_rows(2, maxrows, 1, maxcols) # 返回数据生成器
        for rowdata in gen: # 循环访问
            if str(rowdata[6].value) == id: # # 第七列：UID编码
                return rowdata[6].row
            else:
                continue

    def setRowData(self, rowIndex, data):
        l = len(data)
        for c in range(l):
            self.ws.cell(rowIndex, c+1, data[c])
    
    def setHeaderStyle(self, wbname, tableHeadline):
        self.loadSheet(wbname)
        for c in range(15):
            column = self.colindex[c]
            pos = '{0}{1}'.format(column, 1)
            self.ws[pos] = tableHeadline[c]
            self.ws[pos].alignment = self.halignment
            self.ws[pos].font = self.hfont
            self.ws[pos].border = self.hborder
            self.ws[pos].fill = self.hfill
            self.ws.column_dimensions[column].width = self.columnWidth[c]
        self.saveSheet()
        self.closeSheet()    
    
    def putDataToCellAndSetStyle(self, row, dataList):
        thresholdInstance = GetSetObj.get(2) # 获取阈值界面对象实例
        for col in range(15):
            pos = '{0}{1}'.format(self.colindex[col], row) # 表格索引 
            self.ws[pos] = dataList[col]
            val = self.ws[pos].value
            if col == 2: # 漏电流
                if val != '-' and float(val) > float(thresholdInstance.paraDict['th_DrainCurrent_Up']):
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_DrainCurrent_Up']):
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 3: # 工作电流
                if val != '-' and float(val) > float(thresholdInstance.paraDict['th_WorkCurrent_Up']):
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_WorkCurrent_Up']):
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 4: # ID核对
                if val != '-' and val == '失败':
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and val == '成功':
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 5: # 在线检测
                if val != '-' and val == '离线':
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and val == '在线':
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 7: # 被测模块引爆电流
                if val != '-' and float(val) > float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 9: # 被测模块引爆电流判断
                if val != '-' and val == '超限':
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and val == '正常':
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 11: # 内置模块引爆电流
                if val != '-' and float(val) > float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 13: # 内置模块引爆电流判断
                if val != '-' and val == '超限':
                    self.ws[pos].style = 'abnormalContentStyle'
                elif val != '-' and val == '正常':
                    self.ws[pos].style = 'normalContentStyle'
                elif val == '-':
                    self.ws[pos].style = 'defaultContentStyle'
            elif col == 14: # 结论
                if val != '-' and val == '失败':
                    self.ws[pos].style = 'resultFailStyle'
                elif val != '-' and val == '通过':
                    self.ws[pos].style = 'resultPassStyle'
            else:
                self.ws[pos].style = 'defaultContentStyle'

    def updateRowDataByUID(self, dataList): # 根据输入UID编码新增行数据或者替换其对应所在行的数据
        maxrows = self.ws.max_row # 每次执行行数据填入都获取最新的最大行（包括表头所在行）
        maxcols = self.ws.max_column # 每次执行行数据填入都获取最新的最大列
        nonDupCnt = 0
        if maxrows >= 2: # 已填入了表头和数据
            gen = self.ws.iter_rows(2, maxrows, 1, maxcols) # 返回数据生成器
            for rowdata in gen: # 循环访问，判断并更新重复检测结果
                if str(rowdata[6].value) == dataList[6]: # # 第六列：UID编码
                    self.putDataToCellAndSetStyle(rowdata[6].row, dataList)
                    break
                else:
                    nonDupCnt = nonDupCnt + 1
                    continue
            if nonDupCnt == maxrows - 1: # 没有重复结果则新添加一行即可
                self.putDataToCellAndSetStyle(maxrows + 1, dataList)
        else: # 只填入了表头，还未填入有数据，直接写入数据到表头下面一行
            self.putDataToCellAndSetStyle(2, dataList)

    def setCellStyle(self, wb):
        defaultColumnStyle = [0, 1, 6, 8, 10, 12] # 0-based start
        markedColumnStyle  = [2, 3, 4, 5, 7, 9, 11, 13, 14] # 0-based start
        thresholdInstance = GetSetObj.get(2) # 获取阈值界面对象实例
        self.loadSheet(wb)
        maxrow = self.ws.max_row + 1
        # 格式化默认数据格式
        for c in defaultColumnStyle:
            for r in range(2, maxrow, 1):
                pos = '{0}{1}'.format(self.colindex[c], r) # 表格索引
                self.ws[pos].style = 'defaultContentStyle'
        # 标记关键数据格式
        for col in markedColumnStyle:
            for r in range(2, maxrow, 1):
                pos = '{0}{1}'.format(self.colindex[col], r) # 表格索引 
                val = self.ws[pos].value
                if col == 2: # 漏电流
                    if val != '-' and float(val) > float(thresholdInstance.paraDict['th_DrainCurrent_Up']):
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_DrainCurrent_Up']):
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 3: # 工作电流
                    if val != '-' and float(val) > float(thresholdInstance.paraDict['th_WorkCurrent_Up']):
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_WorkCurrent_Up']):
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 4: # ID核对
                    if val != '-' and val == '失败':
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and val == '成功':
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 5: # 在线检测
                    if val != '-' and val == '离线':
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and val == '在线':
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 7: # 被测模块引爆电流
                    if val != '-' and float(val) > float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 9: # 被测模块引爆电流判断
                    if val != '-' and val == '超限':
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and val == '正常':
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 11: # 内置模块引爆电流
                    if val != '-' and float(val) > float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and float(val) <= float(thresholdInstance.paraDict['th_FireCurrent_Up']):
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 13: # 内置模块引爆电流判断
                    if val != '-' and val == '超限':
                        self.ws[pos].style = 'abnormalContentStyle'
                    elif val != '-' and val == '正常':
                        self.ws[pos].style = 'normalContentStyle'
                    elif val == '-':
                        self.ws[pos].style = 'defaultContentStyle'
                elif col == 14: # 结论
                    if val != '-' and val == '失败':
                        self.ws[pos].style = 'resultFailStyle'
                    elif val != '-' and val == '通过':
                        self.ws[pos].style = 'resultPassStyle'
        self.saveSheet()
        self.closeSheet()